"""Parse tracking_packing_slips.xlsx for bulk import."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

EXCEL_TYPE_TO_ORDER_TYPE = {
    "installation": "installation",
    "delivery": "delivery",
    "shipping": "shipment",
    "will call": "will_call",
}

USER_COLUMNS = ("MC", "IS", "SM", "GR", "OO", "RO")
DATE_COLUMNS = tuple(f"date{col}" for col in USER_COLUMNS) + ("datePA", "date_created")


@dataclass
class TrackingSlipRow:
    row_number: int
    slip: str
    customer: str | None
    order_type: str
    notes: str | None
    is_completed: str | None
    is_paid: bool
    is_invoiced: bool
    date_created: datetime | None
    MC: str | None = None
    IS: str | None = None
    SM: str | None = None
    GR: str | None = None
    OO: str | None = None
    RO: str | None = None
    PA: str | None = None
    dateMC: datetime | None = None
    dateIS: datetime | None = None
    dateSM: datetime | None = None
    dateGR: datetime | None = None
    dateOO: datetime | None = None
    dateRO: datetime | None = None
    datePA: datetime | None = None

    @property
    def is_backordered(self) -> bool:
        return (self.is_completed or "").strip().lower() == "back order"

    @property
    def is_order_completed(self) -> bool:
        return (self.is_completed or "").strip().lower() == "completed"

    def should_skip(self) -> str | None:
        if not self.slip or not str(self.slip).strip():
            return "missing slip number"
        notes_lower = (self.notes or "").lower()
        if "delete" in notes_lower or "void" in notes_lower:
            return "notes contain delete/void"
        return None


def _to_utc_datetime(value) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, datetime.min.time())
    elif isinstance(value, (int, float)):
        # Excel serial date fallback
        from openpyxl.utils.datetime import from_excel
        dt = from_excel(value)
    else:
        return None

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def _parse_is_paid(value) -> bool:
    if value is None:
        return False
    text = str(value).strip().upper().replace(" ", "_")
    return text not in ("", "NOT_PAID", "NOTPAID", "FALSE", "0", "NO")


def _parse_is_invoiced(pa_value) -> bool:
    if pa_value is None:
        return False
    text = str(pa_value).strip().lower()
    return text in ("invoiced", "received")


def _map_order_type(excel_type: str | None) -> str:
    if not excel_type:
        return "installation"
    mapped = EXCEL_TYPE_TO_ORDER_TYPE.get(str(excel_type).strip().lower())
    return mapped or "installation"


def _normalize_slip(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _row_from_dict(row_number: int, data: dict) -> TrackingSlipRow:
    pa = data.get("PA")
    return TrackingSlipRow(
        row_number=row_number,
        slip=_normalize_slip(data.get("slip")),
        customer=(str(data["customer"]).strip() if data.get("customer") else None),
        order_type=_map_order_type(data.get("type")),
        notes=(str(data["notes"]).strip() if data.get("notes") else None),
        is_completed=(str(data["is_completed"]).strip() if data.get("is_completed") else None),
        is_paid=_parse_is_paid(data.get("is_paid")),
        is_invoiced=_parse_is_invoiced(pa),
        date_created=_to_utc_datetime(data.get("date_created")),
        MC=data.get("MC"),
        IS=data.get("IS"),
        SM=data.get("SM"),
        GR=data.get("GR"),
        OO=data.get("OO"),
        RO=data.get("RO"),
        PA=pa,
        dateMC=_to_utc_datetime(data.get("dateMC")),
        dateIS=_to_utc_datetime(data.get("dateIS")),
        dateSM=_to_utc_datetime(data.get("dateSM")),
        dateGR=_to_utc_datetime(data.get("dateGR")),
        dateOO=_to_utc_datetime(data.get("dateOO")),
        dateRO=_to_utc_datetime(data.get("dateRO")),
        datePA=_to_utc_datetime(data.get("datePA")),
    )


def iter_tracking_slip_rows(
    xlsx_path: str | Path,
    *,
    slip_filter: str | None = None,
) -> Iterator[TrackingSlipRow]:
    path = Path(xlsx_path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]

        for row_number, raw_row in enumerate(rows_iter, start=2):
            data = {
                headers[i]: raw_row[i] if i < len(raw_row) else None
                for i in range(len(headers))
                if headers[i]
            }
            row = _row_from_dict(row_number, data)
            if slip_filter is not None and row.slip != str(slip_filter).strip():
                continue
            yield row
    finally:
        wb.close()
