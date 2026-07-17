"""Parse Item List.xlsm sheet Additions for new air filter imports."""

from __future__ import annotations

import re
from dataclasses import dataclass
from fractions import Fraction
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

ITEM_COL = 2
DESCRIPTION_COL = 4
TYPE_COL = 6
QOH_COL = 8
VENDOR_COL = 10


@dataclass
class ItemListAdditionRow:
    row_number: int
    item: str
    description: str | None
    filter_type: str | None
    quantity_on_hand: int
    preferred_vendor: str | None
    height: int
    width: int
    depth: int
    merv_rating: int

    def should_skip(self) -> str | None:
        if not self.item:
            return "missing item"
        if self.height <= 0 or self.width <= 0:
            return "could not parse dimensions from description"
        return None


def _parse_int(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return 0


def _parse_dim_token(token: str) -> int:
    token = token.strip().strip(";").strip()
    if not token:
        raise ValueError("empty dimension token")

    parts = token.split()
    total = 0.0
    i = 0
    while i < len(parts):
        part = parts[i]
        if "/" in part and " " not in part:
            total += float(Fraction(part))
            i += 1
        elif i + 1 < len(parts) and "/" in parts[i + 1]:
            total += float(int(parts[i])) + float(Fraction(parts[i + 1]))
            i += 2
        else:
            total += float(part)
            i += 1
    return int(round(total))


def parse_dimensions(description: str | None) -> tuple[int, int, int] | None:
    if not description:
        return None

    text = str(description).strip()
    dim_token = r"\d+(?:\s+\d+/\d+|\.\d+|/\d+)?"
    dim_end = r"(?=\s|$|[^0-9./\s])"
    three_dim = re.match(
        rf"^({dim_token})\s*x\s*({dim_token})\s*x\s*({dim_token})",
        text,
        re.IGNORECASE,
    )
    if three_dim:
        return (
            _parse_dim_token(three_dim.group(1)),
            _parse_dim_token(three_dim.group(2)),
            _parse_dim_token(three_dim.group(3)),
        )

    two_dim = re.match(
        rf"^({dim_token})\s*x\s*({dim_token}(?:\s+\d+/\d+)?){dim_end}",
        text,
        re.IGNORECASE,
    )
    if not two_dim:
        two_dim = re.search(
            rf"({dim_token})\s*x\s*({dim_token}(?:\s+\d+/\d+)?){dim_end}",
            text,
            re.IGNORECASE,
        )
    if two_dim:
        return (
            _parse_dim_token(two_dim.group(1)),
            _parse_dim_token(two_dim.group(2)),
            0,
        )

    return None


def parse_merv_rating(description: str | None, *, default: int = 10) -> int:
    if not description:
        return default
    match = re.search(r"MERV\s*(\d+)", str(description), re.IGNORECASE)
    if match:
        return int(match.group(1))
    return default


def iter_item_list_addition_rows(
    xlsx_path: Path,
    *,
    item_filter: str | None = None,
) -> Iterator[ItemListAdditionRow]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["Additions"]
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue

            item_raw = row[ITEM_COL] if len(row) > ITEM_COL else None
            if item_raw is None or str(item_raw).strip() == "":
                continue

            item = str(item_raw).strip()
            if item_filter is not None and item.lower() != item_filter.strip().lower():
                continue

            description = row[DESCRIPTION_COL] if len(row) > DESCRIPTION_COL else None
            if description is not None:
                description = str(description).strip() or None

            filter_type = row[TYPE_COL] if len(row) > TYPE_COL else None
            if filter_type is not None:
                filter_type = str(filter_type).strip() or None

            preferred_vendor = row[VENDOR_COL] if len(row) > VENDOR_COL else None
            if preferred_vendor is not None:
                preferred_vendor = str(preferred_vendor).strip() or None

            dims = parse_dimensions(description)
            height, width, depth = dims if dims else (0, 0, 0)

            yield ItemListAdditionRow(
                row_number=row_number,
                item=item,
                description=description,
                filter_type=filter_type,
                quantity_on_hand=_parse_int(row[QOH_COL] if len(row) > QOH_COL else None),
                preferred_vendor=preferred_vendor,
                height=height,
                width=width,
                depth=depth,
                merv_rating=parse_merv_rating(description),
            )
    finally:
        wb.close()
