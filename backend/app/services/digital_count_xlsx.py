"""Parse Digital Count.xlsm for warehouse physical-count imports."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

ITEM_COL = 2
QOH_COL = 8
PHYSICAL_COUNT_COL = 10


@dataclass
class DigitalCountRow:
    row_number: int
    item: str
    description: str | None
    quantity_on_hand: int | None
    physical_count: int

    def should_skip(self) -> str | None:
        if not self.item or not str(self.item).strip():
            return "missing item"
        if self.physical_count <= 0:
            return "physical count is zero or negative"
        return None


def _parse_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def iter_digital_count_rows(
    xlsx_path: Path,
    *,
    item_filter: str | None = None,
) -> Iterator[DigitalCountRow]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["Sheet1"]
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue

            item_raw = row[ITEM_COL] if len(row) > ITEM_COL else None
            if item_raw is None or str(item_raw).strip() == "":
                continue

            item = str(item_raw).strip()
            if item_filter is not None and item.lower() != item_filter.strip().lower():
                continue

            description = row[4] if len(row) > 4 else None
            if description is not None:
                description = str(description).strip() or None

            quantity_on_hand = _parse_int(row[QOH_COL] if len(row) > QOH_COL else None)
            physical_count = _parse_int(row[PHYSICAL_COUNT_COL] if len(row) > PHYSICAL_COUNT_COL else None)
            if physical_count is None:
                physical_count = 0

            yield DigitalCountRow(
                row_number=row_number,
                item=item,
                description=description,
                quantity_on_hand=quantity_on_hand,
                physical_count=physical_count,
            )
    finally:
        wb.close()
